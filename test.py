import openpyxl
wb =openpyxl.load_workbook(filename='aa.xlsx')
finalSheet= wb.active

userList=[]
for i in range(4,100):
    emptyDic = {
        'username':'',
        'firstName': '',
        'lastName': '',
        'email': ''
    }
    email=finalSheet.cell(row=i, column=7).value
    emptyDic['email']=email

    userName=finalSheet.cell(row=i, column=6).value
    emptyDic['username']=userName

    fullName=finalSheet.cell(row=i, column=5).value
    emptyDic['firstName']=fullName.split(' ')[0]

    lastName=fullName.split(' ')[1]
    if len(fullName.split()[0:]) >2:
        for i in range(2, len(fullName.split()[0:])):
            lastName +=' '+ fullName.split()[i]
    emptyDic['lastName']=lastName

    userList.append(emptyDic)

for userList in userList:
    print(userList)